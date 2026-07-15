"""
Builds Permian_Crude_Hedge_Model.xlsx -- a student/intern-level educational
workbook comparing three hedge strategies for a hypothetical Permian Basin
crude producer against real EIA WTI Cushing spot price history.

Run:  python build_model.py
"""

import datetime as dt

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

import data_loader as dl

OUTPUT_PATH = "Permian_Crude_Hedge_Model.xlsx"
ANALYSIS_START = "2023-07-01"
ANALYSIS_END = "2026-06-01"

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
FONT_NAME = "Arial"
TITLE = Font(name=FONT_NAME, size=16, bold=True, color="1F3864")
SUBTITLE = Font(name=FONT_NAME, size=10, italic=True, color="595959")
SECTION = Font(name=FONT_NAME, size=12, bold=True, color="1F3864")
HEADER = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
LABEL = Font(name=FONT_NAME, size=10, bold=True)
BODY = Font(name=FONT_NAME, size=10)
NOTE = Font(name=FONT_NAME, size=9, italic=True, color="7F7F7F")
INPUT_FONT = Font(name=FONT_NAME, size=10, bold=True, color="0000FF")
LINK_FONT = Font(name=FONT_NAME, size=10, color="007A33")

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
SUMMARY_FILL = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_PRICE = '$#,##0.00'
FMT_USD = '$#,##0;($#,##0)'
FMT_PCT = '0.0%'
FMT_BBL = '#,##0'
FMT_DATE = 'mmm-yyyy'

SHEET_INPUTS = "Inputs & Assumptions"
SHEET_HIST = "Historical WTI Data"
SHEET_CALC = "Hedge Calculations"
SHEET_STRESS = "Stress Scenarios"
SHEET_SUMMARY = "Executive Summary"


def style_title(ws, cell, text, span=6):
    ws[cell] = text
    ws[cell].font = TITLE
    ws.merge_cells(start_row=ws[cell].row, start_column=ws[cell].column,
                    end_row=ws[cell].row, end_column=ws[cell].column + span - 1)


def style_subtitle(ws, cell, text, span=6):
    ws[cell] = text
    ws[cell].font = SUBTITLE
    ws.merge_cells(start_row=ws[cell].row, start_column=ws[cell].column,
                    end_row=ws[cell].row, end_column=ws[cell].column + span - 1)


def style_section(ws, cell, text, span=6):
    r, c = ws[cell].row, ws[cell].column
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)
    ws[cell] = text
    ws[cell].font = SECTION
    ws[cell].fill = SECTION_FILL
    for col in range(c, c + span):
        ws.cell(row=r, column=col).fill = SECTION_FILL


def header_row(ws, row, col_start, headers):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Sheet 2: Inputs & Assumptions (built first -- other sheets reference it)
# ---------------------------------------------------------------------------
def build_inputs_sheet(wb):
    ws = wb.create_sheet(SHEET_INPUTS)
    set_widths(ws, {"A": 40, "B": 16, "C": 12, "D": 46})

    style_title(ws, "A1", "Inputs & Assumptions", span=4)
    style_subtitle(
        ws, "A2",
        "Hypothetical Permian Basin producer. All figures below are illustrative "
        "model assumptions, not current market quotes or a real hedge program.",
        span=4,
    )

    style_section(ws, "A4", "Illustrative Assumptions (edit the highlighted cells)", span=4)
    header_row(ws, 5, 1, ["Assumption", "Value", "Units", "Notes"])

    rows = [
        ("Monthly forecast production", 50000, "bbl/mo",
         "Planned monthly production for the hypothetical producer.", FMT_BBL),
        ("Actual production (% of forecast)", 0.92, "%",
         "Illustrative production variance vs. forecast.", FMT_PCT),
        ("Hedge percentage", 0.85, "%",
         "Share of forecast production hedged with the WTI swap.", FMT_PCT),
        ("Fixed WTI swap price", 75.00, "$/bbl",
         "Illustrative fixed price on the WTI swap.", FMT_PRICE),
        ("Assumed Midland-Cushing basis", -1.50, "$/bbl",
         "Illustrative differential; not a live market quote.", FMT_PRICE),
        ("Locked Midland-Cushing basis level", -1.50, "$/bbl",
         "Basis level locked in by the basis hedge.", FMT_PRICE),
        ("Analysis start date", dt.date(2023, 7, 1), "date",
         "First month of the hedge comparison.", FMD := 'mmm-yyyy'),
        ("Analysis end date", dt.date(2026, 6, 1), "date",
         "Last month of the hedge comparison.", FMD),
    ]
    first_input_row = 6
    for i, (label, value, units, note, fmt) in enumerate(rows):
        r = first_input_row + i
        ws.cell(row=r, column=1, value=label).font = BODY
        vcell = ws.cell(row=r, column=2, value=value)
        vcell.font = INPUT_FONT
        vcell.fill = INPUT_FILL
        vcell.number_format = fmt
        vcell.border = BOX
        ws.cell(row=r, column=3, value=units).font = NOTE
        ws.cell(row=r, column=4, value=note).font = NOTE
        for c in (1, 3, 4):
            ws.cell(row=r, column=c).border = BOX

    # Data validation on the editable cells
    dv_pct = DataValidation(type="decimal", operator="between", formula1=0, formula2=2)
    dv_pct.error = "Enter a percentage between 0% and 200%."
    ws.add_data_validation(dv_pct)
    dv_pct.add(ws["B7"])

    dv_hedge = DataValidation(type="decimal", operator="between", formula1=0, formula2=1)
    dv_hedge.error = "Hedge percentage must be between 0% and 100%."
    ws.add_data_validation(dv_hedge)
    dv_hedge.add(ws["B8"])

    ws["B10"].comment = Comment(
        "Midland-Cushing basis is a model assumption, not a market quote. "
        "A more-negative basis means Midland crude sells below Cushing WTI.",
        "Model")
    ws["B8"].comment = Comment(
        "Hedging less than 100% of forecast production leaves some volume "
        "unhedged and reduces the risk of over-hedging if production falls short.",
        "Model")

    style_section(ws, "A15", "Observed Data", span=4)
    header_row(ws, 16, 1, ["Item", "Value", "Units", "Notes"])
    ws["A17"] = "Historical WTI Cushing price"
    ws["A17"].font = BODY
    ws["B17"] = "See 'Historical WTI Data' tab"
    ws["B17"].font = LINK_FONT
    ws["C17"] = "$/bbl"
    ws["C17"].font = NOTE
    ws["D17"] = "Real monthly-average data published by the EIA."
    ws["D17"].font = NOTE

    style_section(ws, "A19", "Calculated Outputs (see 'Hedge Calculations' tab)", span=4)
    header_row(ws, 20, 1, ["Output", "Value", "Units", "Notes"])
    calc_rows = [
        ("Avg. net realized price, combined strategy", f"='{SHEET_CALC}'!D6", "$/bbl",
         "Average across the analysis window."),
        ("Total hedge P&L, combined strategy", f"='{SHEET_CALC}'!D10", "$",
         "WTI swap P&L plus basis hedge P&L, summed."),
        ("Total over-hedged volume", f"='{SHEET_CALC}'!D11", "bbl",
         "Hedged volume exceeding actual production, summed."),
    ]
    for i, (label, formula, units, note) in enumerate(calc_rows):
        r = 21 + i
        ws.cell(row=r, column=1, value=label).font = BODY
        vcell = ws.cell(row=r, column=2, value=formula)
        vcell.font = LINK_FONT
        vcell.number_format = FMT_PRICE if "price" in label else (FMT_USD if "P&L" in label else FMT_BBL)
        ws.cell(row=r, column=3, value=units).font = NOTE
        ws.cell(row=r, column=4, value=note).font = NOTE

    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False
    return {"first_input_row": first_input_row}


# ---------------------------------------------------------------------------
# Sheet 3: Historical WTI Data
# ---------------------------------------------------------------------------
def build_historical_sheet(wb, wti_df, meta):
    ws = wb.create_sheet(SHEET_HIST)
    set_widths(ws, {"A": 28, "B": 24, "C": 34, "D": 20})

    style_title(ws, "A1", "Historical WTI Data", span=4)
    style_subtitle(ws, "A2", "Source data used to drive the hedge model below.", span=4)

    meta_rows = [
        ("Series name:", meta["series_name"]),
        ("Source:", meta["source"]),
        ("Units:", meta["units"]),
        ("Original frequency:", meta["original_frequency"]),
        ("Aggregation method:", "EIA daily spot prices aggregated to a simple monthly average."),
        ("Frequency used in workbook:", meta["frequency_used"]),
        ("Date range:", meta["date_range"]),
        ("Retrieval date:", meta["retrieval_date"]),
        ("Source citation:", f"{meta['source']} - {meta['source_url']}"),
    ]
    start = 4
    for i, (label, value) in enumerate(meta_rows):
        r = start + i
        ws.cell(row=r, column=1, value=label).font = LABEL
        vc = ws.cell(row=r, column=2, value=value)
        vc.font = BODY
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

    header_r = start + len(meta_rows) + 1  # row 14
    header_row(ws, header_r, 1, ["Month", "WTI Cushing Price ($/bbl)", "Source", "Notes"])

    data_start = header_r + 1
    for i, row in wti_df.iterrows():
        r = data_start + i
        dcell = ws.cell(row=r, column=1, value=row["Month"].date())
        dcell.number_format = FMT_DATE
        pcell = ws.cell(row=r, column=2, value=float(row["WTI_Cushing_Price_USD_per_BBL"]))
        pcell.number_format = FMT_PRICE
        ws.cell(row=r, column=3, value=row["Source"]).font = NOTE
        ws.cell(row=r, column=4, value=row["Notes"] if pd.notna(row["Notes"]) else "").font = NOTE
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BOX
    data_end = data_start + len(wti_df) - 1

    ws.auto_filter.ref = f"A{header_r}:D{data_end}"
    ws.freeze_panes = f"A{data_start}"
    ws.sheet_view.showGridLines = False

    return {"header_row": header_r, "data_start": data_start, "data_end": data_end}


# ---------------------------------------------------------------------------
# Sheet 4: Hedge Calculations
# ---------------------------------------------------------------------------
def build_hedge_calc_sheet(wb, hist_info, n_months):
    ws = wb.create_sheet(SHEET_CALC)
    widths = {get_column_letter(i): 15 for i in range(1, 19)}
    widths["A"] = 22
    set_widths(ws, widths)

    style_title(ws, "A1", "Hedge Calculations", span=8)
    style_subtitle(
        ws, "A2",
        "Monthly, formula-driven comparison of the three strategies over the analysis window.",
        span=8,
    )

    style_section(ws, "A4", "Summary Metrics", span=4)
    header_row(ws, 5, 1, ["Metric", "Unhedged", "WTI Swap", "Combined (Swap + Basis Hedge)"])

    table_header_row = 13
    data_start = table_header_row + 1
    data_end = data_start + n_months - 1

    summary_defs = [
        ("Average net realized price ($/bbl)", "P", "Q", "R", "AVERAGE", FMT_PRICE),
        ("Average monthly net revenue ($)", "K", "M", "O", "AVERAGE", FMT_USD),
        ("Minimum monthly net revenue ($)", "K", "M", "O", "MIN", FMT_USD),
        ("Maximum monthly net revenue ($)", "K", "M", "O", "MAX", FMT_USD),
    ]
    for i, (label, uc, sc, cc, fn, fmt) in enumerate(summary_defs):
        r = 6 + i
        lcell = ws.cell(row=r, column=1, value=label)
        lcell.font = BODY
        lcell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=r, column=2, value=f"={fn}({uc}{data_start}:{uc}{data_end})")
        ws.cell(row=r, column=3, value=f"={fn}({sc}{data_start}:{sc}{data_end})")
        ws.cell(row=r, column=4, value=f"={fn}({cc}{data_start}:{cc}{data_end})")
        for c in (2, 3, 4):
            cell = ws.cell(row=r, column=c)
            cell.number_format = fmt
            cell.fill = SUMMARY_FILL
            cell.border = BOX

    # Total hedge P&L
    r = 10
    lcell = ws.cell(row=r, column=1, value="Total hedge P&L ($)")
    lcell.font = BODY
    lcell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.cell(row=r, column=2, value=0)
    ws.cell(row=r, column=3, value=f"=SUM(L{data_start}:L{data_end})")
    ws.cell(row=r, column=4, value=f"=SUM(L{data_start}:L{data_end})+SUM(N{data_start}:N{data_end})")
    for c in (2, 3, 4):
        cell = ws.cell(row=r, column=c)
        cell.number_format = FMT_USD
        cell.fill = SUMMARY_FILL
        cell.border = BOX

    # Total over-hedged volume
    r = 11
    lcell = ws.cell(row=r, column=1, value="Total over-hedged volume (bbl)")
    lcell.font = BODY
    lcell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.cell(row=r, column=2, value=0)
    ws.cell(row=r, column=3, value=f"=SUM(I{data_start}:I{data_end})")
    ws.cell(row=r, column=4, value=f"=SUM(I{data_start}:I{data_end})")
    for c in (2, 3, 4):
        cell = ws.cell(row=r, column=c)
        cell.number_format = FMT_BBL
        cell.fill = SUMMARY_FILL
        cell.border = BOX

    for r in range(6, 12):
        ws.row_dimensions[r].height = 28

    # Monthly detail table
    headers = [
        "Date", "WTI Cushing Price ($/bbl)", "Assumed Midland Basis ($/bbl)",
        "Midland Physical Price ($/bbl)", "Forecast Production (bbl)", "Actual Production (bbl)",
        "Hedge %", "Hedged Volume (bbl)", "Over-Hedged Volume (bbl)", "Physical Revenue ($)",
        "Unhedged Net Revenue ($)", "WTI Swap P&L ($)", "WTI Swap Net Revenue ($)",
        "Basis Hedge P&L ($)", "Combined Net Revenue ($)", "Unhedged Net Realized Price ($/bbl)",
        "WTI Swap Net Realized Price ($/bbl)", "Combined Net Realized Price ($/bbl)",
    ]
    header_row(ws, table_header_row, 1, headers)

    hist_price_range = f"'{SHEET_HIST}'!$B${hist_info['data_start']}:$B${hist_info['data_end']}"
    hist_month_range = f"'{SHEET_HIST}'!$A${hist_info['data_start']}:$A${hist_info['data_end']}"

    for i in range(n_months):
        r = data_start + i
        if i == 0:
            ws.cell(row=r, column=1, value="=AnalysisStart")
        else:
            ws.cell(row=r, column=1, value=f"=EDATE($A${data_start},{i})")
        ws.cell(row=r, column=2, value=f"=INDEX({hist_price_range},MATCH(A{r},{hist_month_range},0))")
        ws.cell(row=r, column=3, value="=BasisAssumed")
        ws.cell(row=r, column=4, value=f"=B{r}+C{r}")
        ws.cell(row=r, column=5, value="=ProdForecast")
        ws.cell(row=r, column=6, value=f"=E{r}*ActualPct")
        ws.cell(row=r, column=7, value="=HedgePct")
        ws.cell(row=r, column=8, value=f"=E{r}*G{r}")
        ws.cell(row=r, column=9, value=f"=MAX(H{r}-F{r},0)")
        ws.cell(row=r, column=10, value=f"=F{r}*D{r}")
        ws.cell(row=r, column=11, value=f"=J{r}")
        ws.cell(row=r, column=12, value=f"=H{r}*(SwapPrice-B{r})")
        ws.cell(row=r, column=13, value=f"=J{r}+L{r}")
        ws.cell(row=r, column=14, value=f"=H{r}*(BasisLocked-C{r})")
        ws.cell(row=r, column=15, value=f"=J{r}+L{r}+N{r}")
        ws.cell(row=r, column=16, value=f"=IF(F{r}=0,0,K{r}/F{r})")
        ws.cell(row=r, column=17, value=f"=IF(F{r}=0,0,M{r}/F{r})")
        ws.cell(row=r, column=18, value=f"=IF(F{r}=0,0,O{r}/F{r})")

        fmts = [FMT_DATE, FMT_PRICE, FMT_PRICE, FMT_PRICE, FMT_BBL, FMT_BBL, FMT_PCT, FMT_BBL,
                FMT_BBL, FMT_USD, FMT_USD, FMT_USD, FMT_USD, FMT_USD, FMT_USD, FMT_PRICE, FMT_PRICE, FMT_PRICE]
        for c, fmt in enumerate(fmts, start=1):
            cell = ws.cell(row=r, column=c)
            cell.number_format = fmt
            cell.border = BOX

    ws.auto_filter.ref = f"A{table_header_row}:R{data_end}"
    ws.freeze_panes = f"A{data_start}"
    ws.sheet_view.showGridLines = False

    return {
        "table_header_row": table_header_row, "data_start": data_start, "data_end": data_end,
        "summary_row_avg_nrp": 6, "summary_row_avg_rev": 7, "summary_row_min_rev": 8,
        "summary_row_max_rev": 9, "summary_row_total_pnl": 10, "summary_row_overhedge": 11,
    }


# ---------------------------------------------------------------------------
# Sheet 5: Stress Scenarios
# ---------------------------------------------------------------------------
def build_stress_sheet(wb, calc_info):
    ws = wb.create_sheet(SHEET_STRESS)
    set_widths(ws, {"A": 34, "B": 16, "C": 22, "D": 22, "E": 22})

    style_title(ws, "A1", "Stress Scenarios", span=5)
    style_subtitle(
        ws, "A2",
        "Each scenario applies a single shock to the base-case assumptions. "
        "The base case uses the average WTI price actually observed over the analysis window.",
        span=5,
    )

    style_section(ws, "A4", "Base Case Reference Values", span=5)
    ds, de = calc_info["data_start"], calc_info["data_end"]
    ws["A5"] = "Base WTI Cushing price ($/bbl)"
    ws["A5"].font = BODY
    ws["B5"] = f"=AVERAGE('{SHEET_CALC}'!B{ds}:B{de})"
    ws["B5"].number_format = FMT_PRICE
    ws["A6"] = "Base Midland basis ($/bbl)"
    ws["A6"].font = BODY
    ws["B6"] = "=BasisAssumed"
    ws["B6"].number_format = FMT_PRICE
    ws["A7"] = "Base actual production (% of forecast)"
    ws["A7"].font = BODY
    ws["B7"] = "=ActualPct"
    ws["B7"].number_format = FMT_PCT
    for r in (5, 6, 7):
        ws.cell(row=r, column=2).fill = SUMMARY_FILL
        ws.cell(row=r, column=2).border = BOX

    header_r = 9
    headers = ["Metric", "Base Case", "A: WTI Price Decline (-$15/bbl)",
               "B: Midland Basis Weakening (-$5/bbl)", "C: Production Shortfall (80% of Forecast)"]
    header_row(ws, header_r, 1, headers)

    # Column letters: B=Base, C=Scenario A, D=Scenario B, E=Scenario C
    cols = ["B", "C", "D", "E"]
    r = header_r + 1  # 10: WTI price
    ws.cell(row=r, column=1, value="WTI Cushing price ($/bbl)").font = BODY
    ws[f"B{r}"] = "=$B$5"
    ws[f"C{r}"] = "=$B$5-15"
    ws[f"D{r}"] = "=$B$5"
    ws[f"E{r}"] = "=$B$5"
    wti_row = r

    r += 1  # 11: Midland basis
    ws.cell(row=r, column=1, value="Midland basis ($/bbl)").font = BODY
    ws[f"B{r}"] = "=$B$6"
    ws[f"C{r}"] = "=$B$6"
    ws[f"D{r}"] = "=$B$6-5"
    ws[f"E{r}"] = "=$B$6"
    basis_row = r

    r += 1  # 12: Actual production %
    ws.cell(row=r, column=1, value="Actual production (% of forecast)").font = BODY
    ws[f"B{r}"] = "=$B$7"
    ws[f"C{r}"] = "=$B$7"
    ws[f"D{r}"] = "=$B$7"
    ws[f"E{r}"] = 0.80
    actpct_row = r

    r += 1  # 13: Midland physical price
    ws.cell(row=r, column=1, value="Midland physical price ($/bbl)").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"={c}{wti_row}+{c}{basis_row}"
    price_row = r

    r += 1  # 14: Actual production
    ws.cell(row=r, column=1, value="Actual production (bbl)").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"=ProdForecast*{c}{actpct_row}"
    actprod_row = r

    r += 1  # 15: Hedged volume
    ws.cell(row=r, column=1, value="Hedged volume (bbl)").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = "=ProdForecast*HedgePct"
    hedged_row = r

    r += 1  # 16: Over-hedged volume
    ws.cell(row=r, column=1, value="Over-hedged volume (bbl)").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"=MAX({c}{hedged_row}-{c}{actprod_row},0)"
    overhedge_row = r

    r += 1  # 17: Physical revenue
    ws.cell(row=r, column=1, value="Physical revenue ($)").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"={c}{actprod_row}*{c}{price_row}"
    physrev_row = r

    r += 1  # 18: WTI swap P&L
    ws.cell(row=r, column=1, value="WTI swap P&L ($)").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"={c}{hedged_row}*(SwapPrice-{c}{wti_row})"
    swappnl_row = r

    r += 1  # 19: Basis hedge P&L
    ws.cell(row=r, column=1, value="Basis hedge P&L ($)").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"={c}{hedged_row}*(BasisLocked-{c}{basis_row})"
    basispnl_row = r

    r += 1  # 20: Unhedged net revenue
    ws.cell(row=r, column=1, value="Unhedged net revenue ($)").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"={c}{physrev_row}"
    unhedged_rev_row = r

    r += 1  # 21: WTI swap net revenue
    ws.cell(row=r, column=1, value="WTI swap net revenue ($)").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"={c}{physrev_row}+{c}{swappnl_row}"
    swap_rev_row = r

    r += 1  # 22: Combined net revenue
    ws.cell(row=r, column=1, value="Combined net revenue ($)").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"={c}{physrev_row}+{c}{swappnl_row}+{c}{basispnl_row}"
    combined_rev_row = r

    r += 1  # 23-25: net realized prices
    ws.cell(row=r, column=1, value="Unhedged net realized price ($/bbl)").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"=IF({c}{actprod_row}=0,0,{c}{unhedged_rev_row}/{c}{actprod_row})"
    unhedged_nrp_row = r
    r += 1
    ws.cell(row=r, column=1, value="WTI swap net realized price ($/bbl)").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"=IF({c}{actprod_row}=0,0,{c}{swap_rev_row}/{c}{actprod_row})"
    swap_nrp_row = r
    r += 1
    ws.cell(row=r, column=1, value="Combined net realized price ($/bbl)").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"=IF({c}{actprod_row}=0,0,{c}{combined_rev_row}/{c}{actprod_row})"
    combined_nrp_row = r

    r += 2
    style_section(ws, f"A{r}", "Change from Base Case ($)", span=5)
    r += 1
    header_row(ws, r, 1, ["Metric", "Base Case", "A: WTI Decline", "B: Basis Weakening", "C: Production Shortfall"])
    r += 1
    ws.cell(row=r, column=1, value="Change in unhedged net revenue").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"={c}{unhedged_rev_row}-$B${unhedged_rev_row}"
    r += 1
    ws.cell(row=r, column=1, value="Change in WTI swap net revenue").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"={c}{swap_rev_row}-$B${swap_rev_row}"
    r += 1
    ws.cell(row=r, column=1, value="Change in combined net revenue").font = BODY
    for c in cols:
        ws[f"{c}{r}"] = f"={c}{combined_rev_row}-$B${combined_rev_row}"
    change_end = r

    price_fmt_rows = {wti_row, basis_row, price_row, unhedged_nrp_row, swap_nrp_row, combined_nrp_row}
    pct_fmt_rows = {actpct_row}
    bbl_fmt_rows = {actprod_row, hedged_row, overhedge_row}
    for row_i in range(wti_row, change_end + 1):
        for c in cols:
            cell = ws[f"{c}{row_i}"]
            if cell.value in (None, ""):
                continue
            if row_i in price_fmt_rows:
                cell.number_format = FMT_PRICE
            elif row_i in pct_fmt_rows:
                cell.number_format = FMT_PCT
            elif row_i in bbl_fmt_rows:
                cell.number_format = FMT_BBL
            else:
                cell.number_format = FMT_USD
            cell.border = BOX

    note_row = change_end + 2
    notes = [
        "Scenario A (WTI decline): the WTI swap gains on hedged volume as WTI falls, "
        "offsetting most of the benchmark-price loss on the hedged share of production.",
        "Scenario B (basis weakening): the WTI swap alone does not respond to basis, so "
        "unhedged and WTI-swap revenue both fall by the same amount. Only the combined "
        "strategy's basis hedge offsets the widening discount.",
        "Scenario C (production shortfall): hedged volume is fixed to forecast production, "
        "so when actual production falls short, part of the hedge is no longer backed by "
        "physical barrels (over-hedged volume), which can amplify swings in net realized price.",
    ]
    for i, txt in enumerate(notes):
        cell = ws.cell(row=note_row + i, column=1, value=txt)
        cell.font = NOTE
        ws.merge_cells(start_row=note_row + i, start_column=1, end_row=note_row + i, end_column=5)
        ws.row_dimensions[note_row + i].height = 28

    ws.sheet_view.showGridLines = False
    return {
        "unhedged_nrp_row": unhedged_nrp_row, "swap_nrp_row": swap_nrp_row,
        "combined_nrp_row": combined_nrp_row,
    }


# ---------------------------------------------------------------------------
# Sheet 1: Executive Summary
# ---------------------------------------------------------------------------
def build_summary_sheet(wb, calc_info):
    ws = wb.create_sheet(SHEET_SUMMARY, 0)
    set_widths(ws, {"A": 26, "B": 16, "C": 18, "D": 18, "E": 16, "F": 14, "G": 16})

    style_title(ws, "A1", "Permian Crude Producer Hedge & Basis Risk Model", span=7)
    style_subtitle(
        ws, "A2",
        "Educational / student research project. Simplified hedge comparison for a "
        "hypothetical Permian Basin crude producer -- not financial advice.",
        span=7,
    )

    ws["A4"] = "Purpose:"
    ws["A4"].font = LABEL
    ws["B4"] = ("Show how WTI benchmark price risk, Midland-Cushing basis risk, and production "
                "volume uncertainty affect a Permian producer's revenue, and how basic hedge "
                "strategies change that exposure.")
    ws["B4"].font = BODY
    ws["B4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("B4:G4")
    ws.row_dimensions[4].height = 30

    ws["A6"] = "Core question:"
    ws["A6"].font = LABEL
    ws["B6"] = ("How do WTI prices, Midland basis, and production uncertainty affect a Permian "
                "producer's revenue, and how do basic hedge strategies change that exposure?")
    ws["B6"].font = BODY
    ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("B6:G6")
    ws.row_dimensions[6].height = 30

    ws["A8"] = "Disclosure:"
    ws["A8"].font = LABEL
    ws["B8"] = ("The producer, production volumes, hedge percentages, swap price, and basis "
                "levels used here are entirely hypothetical and illustrative. Only the WTI "
                "Cushing spot price history is real, public EIA data.")
    ws["B8"].font = NOTE
    ws["B8"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("B8:G8")
    ws.row_dimensions[8].height = 30

    style_section(ws, "A10", "Strategy Comparison", span=7)
    header_row(ws, 11, 1, ["Strategy", "Avg. Net Realized Price ($/bbl)", "Avg. Monthly Net Revenue ($)",
                            "Worst Monthly Net Revenue ($)", "Total Hedge P&L ($)",
                            "Revenue Range ($)", "Over-Hedged Volume (bbl)"])

    strategies = [
        ("Unhedged", "B"),
        ("WTI Fixed-Price Swap", "C"),
        ("WTI Swap + Midland Basis Hedge", "D"),
    ]
    for i, (name, col) in enumerate(strategies):
        r = 12 + i
        ws.cell(row=r, column=1, value=name).font = BODY
        ws.cell(row=r, column=2, value=f"='{SHEET_CALC}'!{col}6").number_format = FMT_PRICE
        ws.cell(row=r, column=3, value=f"='{SHEET_CALC}'!{col}7").number_format = FMT_USD
        ws.cell(row=r, column=4, value=f"='{SHEET_CALC}'!{col}8").number_format = FMT_USD
        ws.cell(row=r, column=5, value=f"='{SHEET_CALC}'!{col}10").number_format = FMT_USD
        ws.cell(row=r, column=6, value=f"='{SHEET_CALC}'!{col}9-'{SHEET_CALC}'!{col}8").number_format = FMT_USD
        ws.cell(row=r, column=7, value=f"='{SHEET_CALC}'!{col}11").number_format = FMT_BBL
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = BOX

    style_section(ws, "A16", "Key Findings", span=7)
    findings = [
        "The WTI fixed-price swap tracks the benchmark closely, cutting the swing in net "
        "realized price that comes from WTI moving up or down.",
        "Midland basis risk is not addressed by the WTI swap alone; only the combined "
        "strategy responds when the Midland discount widens or narrows.",
        "Because the hedge is sized to forecast production, a production shortfall can "
        "leave part of the hedge unmatched by physical barrels (over-hedged volume).",
    ]
    for i, txt in enumerate(findings):
        cell = ws.cell(row=17 + i, column=1, value=f"- {txt}")
        cell.font = BODY
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=17 + i, start_column=1, end_row=17 + i, end_column=7)
        ws.row_dimensions[17 + i].height = 28

    style_section(ws, "A21", "Main Remaining Risk by Strategy", span=7)
    risks = [
        ("Unhedged", "Full exposure to both WTI price moves and Midland basis moves."),
        ("WTI Fixed-Price Swap", "Midland basis risk remains; a wider discount still lowers realized price."),
        ("WTI Swap + Basis Hedge", "Production-volume risk remains; a shortfall can create over-hedged volume."),
    ]
    header_row(ws, 22, 1, ["Strategy", "Main Remaining Risk"])
    ws.merge_cells("B22:G22")
    for i, (name, risk) in enumerate(risks):
        r = 23 + i
        ws.cell(row=r, column=1, value=name).font = BODY
        ws.cell(row=r, column=2, value=risk).font = BODY
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = BOX

    style_section(ws, "A27", "Conclusion", span=7)
    conclusion = (
        "Under the selected assumptions, the WTI swap reduces benchmark-price exposure, "
        "while the combined swap and basis hedge creates greater revenue stability. "
        "Production uncertainty remains important because hedging forecast volumes can "
        "create over-hedging. This is an educational, illustrative comparison, not a "
        "professional recommendation or an optimal hedge."
    )
    ws["A28"] = conclusion
    ws["A28"].font = BODY
    ws["A28"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A28:G30")

    ws.sheet_view.showGridLines = False

    # --- Charts ---
    ds, de = calc_info["data_start"], calc_info["data_end"]
    cats = Reference(wb[SHEET_CALC], min_col=1, min_row=ds, max_row=de)

    chart1 = LineChart()
    chart1.title = "Monthly Net Realized Price by Strategy"
    chart1.y_axis.title = "Net Realized Price ($/bbl)"
    chart1.x_axis.title = "Month"
    chart1.x_axis.number_format = FMT_DATE
    chart1.height, chart1.width = 9, 17
    for col in (16, 17, 18):  # P, Q, R
        data = Reference(wb[SHEET_CALC], min_col=col, min_row=calc_info["table_header_row"], max_row=de)
        chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    ws.add_chart(chart1, "A32")

    chart2 = LineChart()
    chart2.title = "Monthly Net Revenue by Strategy"
    chart2.y_axis.title = "Net Revenue ($)"
    chart2.x_axis.title = "Month"
    chart2.x_axis.number_format = FMT_DATE
    chart2.height, chart2.width = 9, 17
    for col in (11, 13, 15):  # K, M, O
        data = Reference(wb[SHEET_CALC], min_col=col, min_row=calc_info["table_header_row"], max_row=de)
        chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    ws.add_chart(chart2, "A51")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def add_named_ranges(wb):
    names = {
        "ProdForecast": f"'{SHEET_INPUTS}'!$B$6",
        "ActualPct": f"'{SHEET_INPUTS}'!$B$7",
        "HedgePct": f"'{SHEET_INPUTS}'!$B$8",
        "SwapPrice": f"'{SHEET_INPUTS}'!$B$9",
        "BasisAssumed": f"'{SHEET_INPUTS}'!$B$10",
        "BasisLocked": f"'{SHEET_INPUTS}'!$B$11",
        "AnalysisStart": f"'{SHEET_INPUTS}'!$B$12",
        "AnalysisEnd": f"'{SHEET_INPUTS}'!$B$13",
    }
    for name, ref in names.items():
        wb.defined_names[name] = DefinedName(name, attr_text=ref)


def main():
    dl.build_clean_csv()
    wti_df = dl.load_wti_data()
    meta = dl.get_metadata(wti_df)

    n_months = len(pd.date_range(ANALYSIS_START, ANALYSIS_END, freq="MS"))

    wb = Workbook()
    wb.remove(wb.active)

    build_inputs_sheet(wb)
    hist_info = build_historical_sheet(wb, wti_df, meta)
    add_named_ranges(wb)
    calc_info = build_hedge_calc_sheet(wb, hist_info, n_months)
    build_stress_sheet(wb, calc_info)
    build_summary_sheet(wb, calc_info)

    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = "1F3864"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = False

    wb.save(OUTPUT_PATH)
    print(f"Saved {OUTPUT_PATH} with sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
